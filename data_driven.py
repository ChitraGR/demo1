import time
from selenium import webdriver
#from selenium.webdriver.chrome.service import Service
#from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import openpyxl
#from webdriver_manager.chrome import ChromeDriverManager

options = Options()
options.add_experimental_option("detach", True)

ser_obj = Service(executable_path=ChromeDriverManager().install())
# options.add_experimental_option("detach", True)
# ser_obj = Service("C:\\Users\\chitra.gr\\Documents\\automation\\chromedriver_win32\\chromedriver.exe")
# driver = webdriver.Chrome(service=ser_obj, options=options)
data_file = "C:\\Users\\chitra.gr\\Documents\\automation\\data.xlsx"
work = openpyxl.load_workbook(data_file)
sheet =work["Sheet1"]
rows = sheet.max_row
cols = sheet.max_column
print(rows)
print(cols)
for r in range(1, rows+1):
    for c in range(1, cols+1):
        print(sheet.cell(r, c).value , end=" ")
    print()

